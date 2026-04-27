"""
Build the migraine GWAS gene set per Phase 2 pre-registration.

Primary: 123 lead-SNP-nearest genes from Hautakangas 2022 Supp Table 3a, Locus ID.
Sensitivity 1: FOCUS + TWAS-prioritized union from Supp Table 11c.
Sensitivity 2: union of nearest + eQTL-mapped (Tables 3a + 9).

Output: TSV files in analysis/data/processed/ with per-gene provenance.
"""
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "data" / "raw" / "hautakangas2022_supp.xlsx"
OUT = ROOT / "data" / "processed"
OUT.mkdir(parents=True, exist_ok=True)


def strip_near(g: str) -> str:
    """Drop 'near ' / 'near ' prefix used in Hautakangas Locus ID column."""
    if g is None:
        return None
    g = str(g).strip().replace(" ", " ")
    g = re.sub(r"^near\s+", "", g, flags=re.IGNORECASE)
    return g


def split_gene_field(field) -> list[str]:
    """Hautakangas often comma- or semicolon-separates multiple genes per locus."""
    if field is None:
        return []
    s = str(field).strip()
    if not s or s.upper() in {"NA", "N/A", "NONE", "-"}:
        return []
    parts = re.split(r"[,;/]\s*", s)
    return [p.strip() for p in parts if p.strip()]


def main() -> None:
    wb = openpyxl.load_workbook(RAW, read_only=True, data_only=True)

    # ---------- Primary: Table 3a, Locus ID column (lead-SNP nearest gene) ----------
    ws = wb["Supplementary Table 3a"]
    rows = list(ws.iter_rows(values_only=True))
    # Header is the row containing "Locus number" and "Locus ID".
    hdr_idx = next(
        i for i, r in enumerate(rows) if r and r[0] == "Locus number" and r[1] == "Locus ID"
    )
    header = rows[hdr_idx]
    locus_id_col = header.index("Locus ID")
    rsid_col = header.index("Rsid")
    chr_col = header.index("Chromosome")

    primary = []
    for r in rows[hdr_idx + 1 :]:
        if not r or r[0] is None:
            continue
        locus_no = r[0]
        locus_id_raw = r[locus_id_col]
        rsid = r[rsid_col]
        chrom = r[chr_col]
        gene = strip_near(locus_id_raw)
        if gene:
            primary.append(
                {
                    "gene": gene,
                    "locus_number": locus_no,
                    "lead_snp": rsid,
                    "chromosome": chrom,
                    "raw_locus_id": locus_id_raw,
                    "method": "lead_snp_nearest",
                }
            )

    print(f"Primary set: {len(primary)} entries from {len(rows[hdr_idx + 1 :])} loci rows")

    primary_path = OUT / "migraine_set_primary.tsv"
    with primary_path.open("w") as fh:
        fh.write("gene\tlocus_number\tlead_snp\tchromosome\traw_locus_id\tmethod\n")
        for e in primary:
            fh.write(
                f"{e['gene']}\t{e['locus_number']}\t{e['lead_snp']}\t"
                f"{e['chromosome']}\t{e['raw_locus_id']}\t{e['method']}\n"
            )

    primary_genes = sorted({e["gene"] for e in primary})
    print(f"Primary unique genes: {len(primary_genes)}")

    # ---------- Sensitivity 1: Table 11c FOCUS + TWAS prioritized ----------
    ws = wb["Supplementary Table 11c"]
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx = next(
        i
        for i, r in enumerate(rows)
        if r and r[0] == "LOCUS NUMBER" and "candidate genes (FOCUS)" in (r[4] if len(r) > 4 else "")
    )
    header = rows[hdr_idx]
    focus_col = header.index("candidate genes (FOCUS)")
    twas_col = header.index("candidate genes (S-PrediXcan/COLOC) ")
    locus_no_col = header.index("LOCUS NUMBER")

    sens1 = []
    for r in rows[hdr_idx + 1 :]:
        if not r or r[locus_no_col] is None:
            continue
        locus_no = r[locus_no_col]
        focus_genes = split_gene_field(r[focus_col])
        twas_genes = split_gene_field(r[twas_col])
        for g in focus_genes:
            sens1.append({"gene": g, "locus_number": locus_no, "method": "FOCUS_PIP_gt_0.5"})
        for g in twas_genes:
            sens1.append(
                {"gene": g, "locus_number": locus_no, "method": "S-PrediXcan_COLOC_PPH4_gt_0.5"}
            )

    print(f"Sensitivity 1 (FOCUS+TWAS) entries: {len(sens1)}")

    sens1_path = OUT / "migraine_set_sensitivity_focus_twas.tsv"
    with sens1_path.open("w") as fh:
        fh.write("gene\tlocus_number\tmethod\n")
        for e in sens1:
            fh.write(f"{e['gene']}\t{e['locus_number']}\t{e['method']}\n")

    sens1_genes = sorted({e["gene"] for e in sens1})
    print(f"Sensitivity 1 unique genes: {len(sens1_genes)}")

    # ---------- Sensitivity 2: nearest + eQTL-mapped union (Tables 3a + 9) ----------
    eqtl_genes = set()
    ws = wb["Supplementary Table 9"]
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx = next(
        i for i, r in enumerate(rows) if r and r[0] == "Locus number" and "Gene symbol" in r
    )
    header = rows[hdr_idx]
    gene_col = header.index("Gene symbol")
    for r in rows[hdr_idx + 1 :]:
        if not r or r[0] is None:
            continue
        g = r[gene_col]
        if g and str(g).strip() and str(g).strip().upper() not in {"NA", "N/A"}:
            eqtl_genes.add(str(g).strip())

    sens2_genes = sorted(set(primary_genes) | eqtl_genes)
    print(
        f"Sensitivity 2 (nearest + eQTL): {len(sens2_genes)} genes"
        f" (nearest={len(primary_genes)}, eQTL={len(eqtl_genes)})"
    )

    sens2_path = OUT / "migraine_set_sensitivity_nearest_plus_eqtl.tsv"
    with sens2_path.open("w") as fh:
        fh.write("gene\tsource\n")
        for g in sens2_genes:
            srcs = []
            if g in primary_genes:
                srcs.append("nearest")
            if g in eqtl_genes:
                srcs.append("eQTL")
            fh.write(f"{g}\t{','.join(srcs)}\n")

    # ---------- Provenance ----------
    provenance = {
        "source_paper": "Hautakangas et al. 2022, Nat Genet 54:152-160, doi:10.1038/s41588-021-00990-0",
        "source_file": str(RAW.relative_to(ROOT)),
        "source_file_sha256": "eb13a123dba46453cbbcdfd18727e54e09aa5c27dc60dd6336503e69c175be3d",
        "primary_set": {
            "rule": "Lead-SNP nearest gene (Locus ID column, with 'near ' prefix stripped)",
            "table": "Supplementary Table 3a",
            "n_loci": len(primary),
            "n_unique_genes": len(primary_genes),
            "output": str(primary_path.relative_to(ROOT)),
        },
        "sensitivity_focus_twas": {
            "rule": "Union of FOCUS PIP>0.5 and S-PrediXcan/COLOC PPH4>0.5 prioritized genes",
            "table": "Supplementary Table 11c",
            "n_entries": len(sens1),
            "n_unique_genes": len(sens1_genes),
            "output": str(sens1_path.relative_to(ROOT)),
        },
        "sensitivity_nearest_plus_eqtl": {
            "rule": "Union of lead-SNP nearest (Table 3a) and lead-variant eQTL-mapped genes (Table 9)",
            "tables": ["Supplementary Table 3a", "Supplementary Table 9"],
            "n_unique_genes": len(sens2_genes),
            "output": str(sens2_path.relative_to(ROOT)),
        },
        "hgnc_remap_pending": True,
    }
    with (OUT / "migraine_set_provenance.json").open("w") as fh:
        json.dump(provenance, fh, indent=2)

    print("Wrote provenance to", OUT / "migraine_set_provenance.json")


if __name__ == "__main__":
    main()
